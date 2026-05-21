#!/usr/bin/env python3
"""
Tor Exit Node Fetcher → Excel Exporter
Fetches IPv4 and IPv6 Tor exit nodes from multiple sources,
batches them, and writes each batch to a separate Excel sheet.

Sources:
  1. dan.me.uk  — https://www.dan.me.uk/torlist/
  2. check.torproject.org — https://check.torproject.org/torbulkexitlist

Usage:
  python fetch_tor_exits.py [--source dan|tor|both] [--batch-size 900]
"""

import argparse
import ipaddress
import logging
import math
import sys
import time
from datetime import datetime
from typing import Optional

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# CONFIGURATION  (edit here or override via CLI)
# ──────────────────────────────────────────────
CONFIG = {
    # Sources
    "sources": {
        "dan": {
            "url": "https://www.dan.me.uk/torlist/",
            "description": "dan.me.uk Tor exit list",
        },
        "tor": {
            "url": "https://check.torproject.org/torbulkexitlist",
            "description": "check.torproject.org bulk exit list",
        },
    },
    # Fetch behaviour
    "batch_size": 900,              # IPs per sheet
    "request_timeout": 30,          # seconds per HTTP request
    "retry_max_attempts": 3,        # retries on failure
    "retry_backoff_base": 2,        # exponential back-off base (seconds)
    "retry_backoff_max": 60,        # cap for back-off delay (seconds)
    "inter_request_sleep": 1.0,     # seconds between source requests
    # SSL verification
    # Point ssl_ca_cert at your corporate / proxy CA bundle (.crt or .pem)
    # so requests trusts your org's TLS inspection certificate.
    # Leave as None to use the system CA store.
    # Set ssl_no_verify to True ONLY for local debugging — never in production.
    "ssl_ca_cert": None,            # e.g. "/etc/ssl/certs/my-corp-ca.crt"
    "ssl_no_verify": False,
    # Output
    "output_file": "tor_exit_nodes.xlsx",
    # Logging
    "log_level": "INFO",            # DEBUG | INFO | WARNING | ERROR
}

# ──────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    level=CONFIG["log_level"],
)
log = logging.getLogger(__name__)


# ──────────────────────────────────────────────
# FETCH  (with retry + back-off)
# ──────────────────────────────────────────────
def fetch_url(url: str, cfg: dict) -> Optional[str]:
    """
    GET *url* with timeout, retries, exponential back-off, and SSL verification.

    SSL behaviour (controlled via cfg):
      ssl_ca_cert   : path to a CA bundle (.crt / .pem) — used for corporate
                      TLS-inspection proxies.  None → system CA store.
      ssl_no_verify : if True, disables certificate validation entirely.
                      ⚠  Only for debugging — never in production.
    """
    if cfg.get("ssl_no_verify"):
        ssl_verify = False
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        log.warning("SSL verification DISABLED — do not use in production!")
    elif cfg.get("ssl_ca_cert"):
        ssl_verify = cfg["ssl_ca_cert"]
        log.debug("SSL verify using CA bundle: %s", ssl_verify)
    else:
        ssl_verify = True   # default: system CA store

    attempt = 0
    while attempt < cfg["retry_max_attempts"]:
        attempt += 1
        try:
            log.info("Fetching %s (attempt %d/%d)", url, attempt, cfg["retry_max_attempts"])
            resp = requests.get(
                url,
                timeout=cfg["request_timeout"],
                verify=ssl_verify,
                headers={
                    "User-Agent": "TorExitFetcher/1.0 (security-policy-update)",
                    "Accept": "text/plain",
                },
            )
            resp.raise_for_status()
            log.info("  → %d bytes received (HTTP %s)", len(resp.content), resp.status_code)
            return resp.text
        except requests.exceptions.SSLError as exc:
            log.error("  SSL error: %s", exc)
            log.error(
                "  Tip: if your network uses TLS inspection, pass your corporate "
                "CA certificate with --ca-cert /path/to/corp-ca.crt"
            )
            # SSL errors are not retried — wrong cert will always fail
            return None
        except requests.exceptions.Timeout:
            log.warning("  Timeout on %s", url)
        except requests.exceptions.HTTPError as exc:
            log.warning("  HTTP error: %s", exc)
        except requests.exceptions.ConnectionError as exc:
            log.warning("  Connection error: %s", exc)
        except requests.exceptions.RequestException as exc:
            log.warning("  Request error: %s", exc)

        if attempt < cfg["retry_max_attempts"]:
            delay = min(cfg["retry_backoff_base"] ** attempt, cfg["retry_backoff_max"])
            log.info("  Retrying in %.1f s…", delay)
            time.sleep(delay)

    log.error("All %d attempts failed for %s", cfg["retry_max_attempts"], url)
    return None


# ──────────────────────────────────────────────
# PARSE
# ──────────────────────────────────────────────
def parse_ips(raw_text: str) -> tuple[list[str], list[str]]:
    """
    Extract valid IPv4 and IPv6 addresses from *raw_text*.
    Lines starting with # are skipped.  Duplicates are removed.
    Returns (ipv4_list, ipv6_list).
    """
    ipv4, ipv6 = [], []
    seen: set[str] = set()
    for line in raw_text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        # strip optional CIDR notation for validation, but keep original
        candidate = line.split()[0]   # ignore trailing comments
        try:
            addr = ipaddress.ip_address(candidate)
            if candidate in seen:
                continue
            seen.add(candidate)
            if addr.version == 4:
                ipv4.append(candidate)
            else:
                ipv6.append(candidate)
        except ValueError:
            log.debug("  Skipping non-IP line: %r", candidate)
    return ipv4, ipv6


# ──────────────────────────────────────────────
# EXCEL HELPERS
# ──────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="1F3864")   # dark navy
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHDR_FILL = PatternFill("solid", start_color="2E75B6")   # blue
SUBHDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="DCE6F1")    # light blue stripe
CENTER      = Alignment(horizontal="center", vertical="center")


def _style_cell(cell, value, font=None, fill=None, align=None):
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align


def add_summary_sheet(wb: Workbook, stats: dict, cfg: dict, per_source_stats: list | None = None):
    """Write a Summary sheet with run metadata, dedup breakdown, and IP counts."""
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22

    row = 1
    title = ws.cell(row=row, column=1, value="Tor Exit Node Export — Unique IPs Only")
    title.font = Font(name="Arial", bold=True, size=14, color="1F3864")
    ws.merge_cells(f"A{row}:E{row}")

    row = 2
    ws.cell(row=row, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC").font = Font(name="Arial", italic=True, size=10, color="595959")
    ws.merge_cells(f"A{row}:E{row}")

    row = 4
    for col, h in enumerate(["Setting", "Value"], start=1):
        _style_cell(ws.cell(row=row, column=col), h, HEADER_FONT, HEADER_FILL, CENTER)

    settings = [
        ("Source(s)", stats.get("sources_used", "—")),
        ("Batch size (IPs per sheet)", cfg["batch_size"]),
        ("Request timeout (s)", cfg["request_timeout"]),
        ("Retry attempts", cfg["retry_max_attempts"]),
        ("Back-off base (s)", cfg["retry_backoff_base"]),
        ("Back-off cap (s)", cfg["retry_backoff_max"]),
    ]
    for i, (k, v) in enumerate(settings, start=5):
        ws.cell(row=i, column=1, value=k).font = DATA_FONT
        ws.cell(row=i, column=2, value=str(v)).font = DATA_FONT

    # ── Dedup breakdown table ──
    row = 12
    ws.cell(row=row, column=1, value="Deduplication Breakdown").font = Font(name="Arial", bold=True, size=12, color="1F3864")
    ws.merge_cells(f"A{row}:E{row}")

    row = 13
    hdrs = ["Source", "IPv4 Fetched", "IPv4 Added (unique)", "IPv6 Fetched", "IPv6 Added (unique)"]
    for col, h in enumerate(hdrs, start=1):
        _style_cell(ws.cell(row=row, column=col), h, SUBHDR_FONT, SUBHDR_FILL, CENTER)

    row = 14
    if per_source_stats:
        for s in per_source_stats:
            fill = ALT_FILL if (row % 2 == 0) else None
            vals = [s["label"], s["raw_v4"], s["new_v4"], s["raw_v6"], s["new_v6"]]
            for col, v in enumerate(vals, start=1):
                c = ws.cell(row=row, column=col, value=v)
                c.font = DATA_FONT
                if fill:
                    c.fill = fill
            row += 1

    # totals row
    _style_cell(ws.cell(row=row, column=1), "TOTAL UNIQUE (final Excel)", SUBHDR_FONT, SUBHDR_FILL, CENTER)
    _style_cell(ws.cell(row=row, column=2), "—", SUBHDR_FONT, SUBHDR_FILL, CENTER)
    _style_cell(ws.cell(row=row, column=3), stats.get("total_ipv4", 0), SUBHDR_FONT, SUBHDR_FILL, CENTER)
    _style_cell(ws.cell(row=row, column=4), "—", SUBHDR_FONT, SUBHDR_FILL, CENTER)
    _style_cell(ws.cell(row=row, column=5), stats.get("total_ipv6", 0), SUBHDR_FONT, SUBHDR_FILL, CENTER)
    row += 2

    note = ws.cell(row=row, column=1, value="✓ All IPs in this workbook are globally unique across all sources. No duplicates exist.")
    note.font = Font(name="Arial", bold=True, size=10, color="375623")
    ws.merge_cells(f"A{row}:E{row}")

    ws.sheet_view.showGridLines = False


def create_ip_sheet(
    wb: Workbook,
    sheet_name: str,
    ip_list: list[str],
    ip_version: int,
    batch_num: int,
    total_batches: int,
    source_label: str,
):
    """Write one batch of IPs to a new worksheet."""
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 30

    # Sheet title row
    title_cell = ws.cell(row=1, column=1, value=f"Tor Exit Nodes — IPv{ip_version} — Batch {batch_num}/{total_batches}")
    title_cell.font = Font(name="Arial", bold=True, size=12, color="1F3864")
    ws.merge_cells("A1:D1")
    title_cell.alignment = CENTER

    # Sub-header row
    sub_cell = ws.cell(row=2, column=1, value=f"Source: {source_label}  |  {len(ip_list)} IPs in this batch")
    sub_cell.font = Font(name="Arial", italic=True, size=10, color="595959")
    ws.merge_cells("A2:D2")

    # Column headers
    headers = ["#", f"IPv{ip_version} Address", "Ver", "Source"]
    for col, h in enumerate(headers, start=1):
        _style_cell(ws.cell(row=3, column=col), h, SUBHDR_FONT, SUBHDR_FILL, CENTER)

    # Data rows
    for i, ip in enumerate(ip_list, start=1):
        row = i + 3
        fill = ALT_FILL if i % 2 == 0 else None
        ws.cell(row=row, column=1, value=i).font = DATA_FONT
        if fill:
            ws.cell(row=row, column=1).fill = fill
        ip_cell = ws.cell(row=row, column=2, value=ip)
        ip_cell.font = Font(name="Courier New", size=10)
        if fill:
            ip_cell.fill = fill
        ws.cell(row=row, column=3, value=f"v{ip_version}").font = DATA_FONT
        if fill:
            ws.cell(row=row, column=3).fill = fill
        src_cell = ws.cell(row=row, column=4, value=source_label)
        src_cell.font = DATA_FONT
        if fill:
            src_cell.fill = fill

    # Auto-filter on header row
    ws.auto_filter.ref = f"A3:{get_column_letter(4)}{len(ip_list)+3}"
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
def build_excel(all_ipv4: list[str], all_ipv6: list[str], source_label: str, cfg: dict, per_source_stats: list | None = None):
    batch_size = cfg["batch_size"]
    wb = Workbook()
    wb.remove(wb.active)   # drop default empty sheet

    def batch_and_write(ip_list: list[str], version: int) -> int:
        total = len(ip_list)
        n_batches = max(1, math.ceil(total / batch_size))
        for b in range(n_batches):
            chunk = ip_list[b * batch_size : (b + 1) * batch_size]
            sheet_name = f"IPv{version}_Batch_{b+1:03d}"
            log.info("  Writing sheet %s (%d IPs)", sheet_name, len(chunk))
            create_ip_sheet(wb, sheet_name, chunk, version, b + 1, n_batches, source_label)
        return n_batches

    log.info("Building Excel workbook…")
    ipv4_sheets = batch_and_write(all_ipv4, 4)
    ipv6_sheets = batch_and_write(all_ipv6, 6)

    stats = {
        "sources_used": source_label,
        "total_ipv4": len(all_ipv4),
        "total_ipv6": len(all_ipv6),
        "ipv4_sheets": ipv4_sheets,
        "ipv6_sheets": ipv6_sheets,
    }
    add_summary_sheet(wb, stats, cfg, per_source_stats)

    out = cfg["output_file"]
    wb.save(out)
    log.info("Saved → %s", out)
    return stats


def run(selected_source: str, cfg: dict):
    sources = cfg["sources"]
    source_keys = list(sources.keys()) if selected_source == "both" else [selected_source]

    # Use sets from the start — insertion order preserved via dict trick at the end
    seen_ipv4: set[str] = set()
    seen_ipv6: set[str] = set()
    ordered_ipv4: dict[str, None] = {}   # preserves first-seen order
    ordered_ipv6: dict[str, None] = {}

    source_labels: list[str] = []
    per_source_stats: list[dict] = []

    for idx, key in enumerate(source_keys):
        src = sources[key]
        log.info("─── Fetching from %s (%s) ───", src["description"], src["url"])

        raw = fetch_url(src["url"], cfg)
        if raw is None:
            log.error("Could not retrieve data from %s — skipping.", src["description"])
            continue

        ipv4, ipv6 = parse_ips(raw)
        log.info("  Raw parsed: %d IPv4, %d IPv6", len(ipv4), len(ipv6))

        # Count how many are new vs already seen from a previous source
        new_v4 = sum(1 for ip in ipv4 if ip not in seen_ipv4)
        new_v6 = sum(1 for ip in ipv6 if ip not in seen_ipv6)
        dup_v4 = len(ipv4) - new_v4
        dup_v6 = len(ipv6) - new_v6

        for ip in ipv4:
            if ip not in seen_ipv4:
                seen_ipv4.add(ip)
                ordered_ipv4[ip] = None
        for ip in ipv6:
            if ip not in seen_ipv6:
                seen_ipv6.add(ip)
                ordered_ipv6[ip] = None

        log.info(
            "  Added: %d new IPv4 (%d cross-source dups dropped), "
            "%d new IPv6 (%d cross-source dups dropped)",
            new_v4, dup_v4, new_v6, dup_v6,
        )
        source_labels.append(src["description"])
        per_source_stats.append({
            "label": src["description"],
            "raw_v4": len(ipv4), "raw_v6": len(ipv6),
            "new_v4": new_v4,    "new_v6": new_v6,
            "dup_v4": dup_v4,    "dup_v6": dup_v6,
        })

        if idx < len(source_keys) - 1:
            log.info("Sleeping %.1f s before next source…", cfg["inter_request_sleep"])
            time.sleep(cfg["inter_request_sleep"])

    if not ordered_ipv4 and not ordered_ipv6:
        log.error("No IPs collected from any source. Aborting.")
        sys.exit(1)

    all_ipv4 = list(ordered_ipv4)
    all_ipv6 = list(ordered_ipv6)

    log.info(
        "Final unique pool: %d IPv4, %d IPv6  (total unique: %d)",
        len(all_ipv4), len(all_ipv6), len(all_ipv4) + len(all_ipv6),
    )

    combined_label = " + ".join(source_labels) if source_labels else "unknown"
    stats = build_excel(all_ipv4, all_ipv6, combined_label, cfg, per_source_stats)

    print("\n" + "═" * 60)
    print("  Tor Exit Node Export — Complete")
    print("═" * 60)
    for s in per_source_stats:
        print(f"  [{s['label']}]")
        print(f"    IPv4: {s['raw_v4']:>6,} fetched  →  {s['new_v4']:>6,} unique added  ({s['dup_v4']:,} cross-source dups dropped)")
        print(f"    IPv6: {s['raw_v6']:>6,} fetched  →  {s['new_v6']:>6,} unique added  ({s['dup_v6']:,} cross-source dups dropped)")
    print("  " + "─" * 56)
    print(f"  Final unique IPv4 : {stats['total_ipv4']:,}  →  {stats['ipv4_sheets']} sheet(s)")
    print(f"  Final unique IPv6 : {stats['total_ipv6']:,}  →  {stats['ipv6_sheets']} sheet(s)")
    print(f"  Output            : {cfg['output_file']}")
    print("═" * 60 + "\n")


# ──────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────
def parse_args():
    p = argparse.ArgumentParser(
        description="Fetch Tor exit nodes and export to Excel (batched sheets).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument(
        "--source",
        choices=["dan", "tor", "both"],
        default="both",
        help="Which source(s) to fetch from.",
    )
    p.add_argument(
        "--batch-size",
        type=int,
        default=CONFIG["batch_size"],
        help="Number of IPs per Excel sheet.",
    )
    p.add_argument(
        "--output",
        default=CONFIG["output_file"],
        help="Output .xlsx file path.",
    )
    p.add_argument(
        "--timeout",
        type=int,
        default=CONFIG["request_timeout"],
        help="HTTP request timeout in seconds.",
    )
    p.add_argument(
        "--retries",
        type=int,
        default=CONFIG["retry_max_attempts"],
        help="Max retry attempts per request.",
    )
    p.add_argument(
        "--backoff-base",
        type=float,
        default=CONFIG["retry_backoff_base"],
        help="Exponential back-off base (seconds).",
    )
    p.add_argument(
        "--backoff-max",
        type=float,
        default=CONFIG["retry_backoff_max"],
        help="Maximum back-off delay cap (seconds).",
    )
    p.add_argument(
        "--sleep",
        type=float,
        default=CONFIG["inter_request_sleep"],
        help="Sleep between source requests (seconds).",
    )
    p.add_argument(
        "--log-level",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        default=CONFIG["log_level"],
        help="Logging verbosity.",
    )

    ssl_group = p.add_argument_group(
        "SSL / TLS",
        "Control certificate verification. Use --ca-cert when your network "
        "terminates TLS (e.g. corporate proxy / Zscaler / Palo Alto NGFW).",
    )
    ssl_group.add_argument(
        "--ca-cert",
        metavar="PATH",
        default=CONFIG["ssl_ca_cert"],
        help=(
            "Path to a CA certificate bundle (.crt or .pem) to use for SSL "
            "verification. Typical locations: "
            "/etc/ssl/certs/ca-certificates.crt (Linux), "
            "~/Library/Application Support/... (macOS), "
            "C:\\certs\\corp-ca.crt (Windows)."
        ),
    )
    ssl_group.add_argument(
        "--no-verify",
        action="store_true",
        default=CONFIG["ssl_no_verify"],
        help=(
            "Disable SSL certificate verification entirely. "
            "⚠  INSECURE — only for local debugging, never in production."
        ),
    )
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()

    logging.getLogger().setLevel(args.log_level)

    cfg = CONFIG.copy()
    cfg["batch_size"]           = args.batch_size
    cfg["output_file"]          = args.output
    cfg["request_timeout"]      = args.timeout
    cfg["retry_max_attempts"]   = args.retries
    cfg["retry_backoff_base"]   = args.backoff_base
    cfg["retry_backoff_max"]    = args.backoff_max
    cfg["inter_request_sleep"]  = args.sleep

    # ── SSL validation ──
    if args.no_verify and args.ca_cert:
        print("ERROR: --no-verify and --ca-cert are mutually exclusive.", file=sys.stderr)
        sys.exit(1)

    if args.ca_cert:
        import os
        if not os.path.isfile(args.ca_cert):
            print(f"ERROR: CA certificate file not found: {args.ca_cert!r}", file=sys.stderr)
            sys.exit(1)
        cfg["ssl_ca_cert"]   = args.ca_cert
        cfg["ssl_no_verify"] = False
    elif args.no_verify:
        cfg["ssl_ca_cert"]   = None
        cfg["ssl_no_verify"] = True
    else:
        cfg["ssl_ca_cert"]   = None
        cfg["ssl_no_verify"] = False

    run(args.source, cfg)
