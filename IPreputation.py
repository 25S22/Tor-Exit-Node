#!/usr/bin/env python3
"""
Bulk IP Reputation Checker
==========================

Checks a large list of IPs (e.g. 19,000) against multiple FREE threat-intel
sources and produces a single Excel report.

Strategy (cheapest/fastest sources first, so you burn as few rate-limited
API calls as possible):

  Tier 0   Local static blocklists (Spamhaus DROP/EDROP, FireHOL, Blocklist.de,
           Feodo Tracker, CINS Army). No API key, no rate limit, covers all
           IPs instantly.
  Tier 0.5 AbuseIPDB "blacklist" endpoint (5 calls/day, up to 10,000 IPs per
           call) -> cross-referenced locally, doesn't touch your 1,000/day
           per-IP check quota.
  Tier 1   Shodan InternetDB (no key, effectively no rate limit). Not a
           malicious/clean verdict, but flags exposed services / known CVEs.
  Tier 2   Individual API lookups for whatever is STILL unresolved after
           Tiers 0-1: AbuseIPDB check, VirusTotal, (optional) GreyNoise.
           These respect each vendor's real daily quota and checkpoint their
           progress, so you can stop the script and re-run it tomorrow and
           it picks up exactly where it left off.

Usage:
    export ABUSEIPDB_API_KEY="..."
    export VT_API_KEY="..."
    export GREYNOISE_API_KEY="..."   # optional
    python3 ip_reputation_checker.py --input ips.txt --output report.xlsx

Re-run the same command on subsequent days to continue past rate limits;
progress is stored in checkpoint.json / quota_tracker.json next to the
script.
"""

import argparse
import bisect
import ipaddress
import json
import os
import sys
import time
from datetime import date
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests

try:
    import pandas as pd
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependencies. Run:\n"
          "  pip install requests pandas openpyxl tqdm --break-system-packages")
    sys.exit(1)

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(iterable, **kwargs):
        return iterable

# ==========================================================================
# ============================  CONFIGURATION  ============================
# ==========================================================================
#
#   Paste your free API keys between the quotes below. Leave a key blank
#   ("") to skip that vendor entirely -- the script will just print
#   "skipping" for that tier and keep going with the others.
#
#   Get free keys here:
#     AbuseIPDB  -> https://www.abuseipdb.com/register
#     VirusTotal -> https://www.virustotal.com/gui/join-us
#     GreyNoise  -> https://viz.greynoise.io/account/   (optional)
#
#   SECURITY NOTE: once you paste real keys in here, this file contains a
#   secret. Don't commit it to a public repo or share it around -- treat it
#   like a password file.
#
# --------------------------------------------------------------------------

ABUSEIPDB_API_KEY = "PASTE_YOUR_ABUSEIPDB_KEY_HERE"
VT_API_KEY        = "PASTE_YOUR_VIRUSTOTAL_KEY_HERE"
GREYNOISE_API_KEY = "PASTE_YOUR_GREYNOISE_KEY_HERE"        # optional, can leave blank

#   Set the input/output paths so you can just run:  python3 ip_reputation_checker.py
#   with no arguments at all. Use an absolute path if the file isn't sitting
#   next to this script (Windows example: r"C:\Users\you\Desktop\ips.txt").
#
#   Examples:
#     INPUT_FILE_PATH  = "ips.txt"
#     INPUT_FILE_PATH  = "/home/you/Desktop/ips.txt"
#     INPUT_FILE_PATH  = r"C:\Users\you\Desktop\ips.txt"
#     OUTPUT_FILE_PATH = "/home/you/Desktop/ip_reputation_report.xlsx"

INPUT_FILE_PATH  = "ips.txt"
OUTPUT_FILE_PATH = "ip_reputation_report.xlsx"

#   If you're behind a corporate proxy/firewall that does SSL/TLS inspection,
#   requests will fail with a certificate verification error unless you point
#   this at your organization's CA bundle (a .pem or .crt file -- ask your IT
#   / security team for it, or export it from your browser's trusted root
#   certificates).
#
#   Examples:
#     SSL_CERT_PATH = ""                                   # use default system certs (normal case)
#     SSL_CERT_PATH = "/home/you/certs/corporate-ca.pem"
#     SSL_CERT_PATH = r"C:\certs\corporate-ca.pem"
#
#   SSL_VERIFY_DISABLE is an escape hatch for quick troubleshooting only --
#   it turns verification off completely, which means your traffic could be
#   intercepted without you knowing. Leave it False; only flip it to True
#   temporarily if you're stuck and don't yet have the CA bundle.

SSL_CERT_PATH = ""
SSL_VERIFY_DISABLE = False

# --------------------------------------------------------------------------
# ==========================================================================



def _resolve_key(hardcoded, env_var_name):
    """
    Hardcoded value above wins if you've actually pasted a real key in.
    Otherwise falls back to an environment variable of the same name, so
    both workflows (paste-in-file, or export in shell) work.
    """
    placeholder_markers = ("PASTE_YOUR_", "")
    if hardcoded and not hardcoded.startswith("PASTE_YOUR_"):
        return hardcoded.strip()
    return os.environ.get(env_var_name, "").strip()


ABUSEIPDB_API_KEY = _resolve_key(ABUSEIPDB_API_KEY, "ABUSEIPDB_API_KEY")
VT_API_KEY = _resolve_key(VT_API_KEY, "VT_API_KEY")
GREYNOISE_API_KEY = _resolve_key(GREYNOISE_API_KEY, "GREYNOISE_API_KEY")

# --- SSL/TLS verification setup ------------------------------------------
# Also honor a CUSTOM_CA_BUNDLE / SSL_CERT_PATH env var, mirroring the API
# key pattern, in case you'd rather not hardcode the path in the file.
_env_cert_path = os.environ.get("SSL_CERT_PATH", "").strip()
if SSL_CERT_PATH.strip():
    _resolved_cert_path = SSL_CERT_PATH.strip()
elif _env_cert_path:
    _resolved_cert_path = _env_cert_path
else:
    _resolved_cert_path = ""

if SSL_VERIFY_DISABLE:
    REQUESTS_VERIFY = False
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    print("WARNING: SSL_VERIFY_DISABLE is True -- certificate verification "
          "is OFF for all requests. This is insecure; only use it temporarily.")
elif _resolved_cert_path:
    if not Path(_resolved_cert_path).exists():
        print(f"WARNING: SSL_CERT_PATH is set to '{_resolved_cert_path}' but "
              f"that file doesn't exist. Falling back to default certificate "
              f"verification, which will likely fail if you're behind an "
              f"SSL-inspecting proxy.")
        REQUESTS_VERIFY = True
    else:
        REQUESTS_VERIFY = _resolved_cert_path
else:
    REQUESTS_VERIFY = True

# Every HTTP call in this script goes through this one session, so the SSL
# setting above is applied consistently everywhere (blocklists, AbuseIPDB,
# VirusTotal, GreyNoise, Shodan InternetDB).
HTTP = requests.Session()
HTTP.verify = REQUESTS_VERIFY

BASE_DIR = Path(__file__).resolve().parent
CHECKPOINT_FILE = BASE_DIR / "checkpoint.json"
QUOTA_FILE = BASE_DIR / "quota_tracker.json"
BLOCKLIST_CACHE_DIR = BASE_DIR / "blocklist_cache"
BLOCKLIST_MAX_AGE_HOURS = 12

# Keep a small buffer under each vendor's published cap so you never get a
# hard 429 mid-run.
DAILY_QUOTAS = {
    "abuseipdb_check": 950,       # published cap: 1,000/day
    "abuseipdb_blacklist": 4,     # published cap: 5/day
    "virustotal": 480,            # published cap: 500/day (also 4/min)
    "greynoise": 900,             # published cap: 1,000/day (free registered key)
}

# Free, no-key, no-rate-limit blocklists (plain-text IP or CIDR per line,
# some with comments starting with ';' or '#').
BLOCKLIST_SOURCES = {
    "spamhaus_drop": "https://www.spamhaus.org/drop/drop.txt",
    "spamhaus_edrop": "https://www.spamhaus.org/drop/edrop.txt",
    "firehol_level1": "https://raw.githubusercontent.com/firehol/blocklist-ipsets/master/firehol_level1.netset",
    "firehol_level2": "https://raw.githubusercontent.com/firehol/blocklist-ipsets/master/firehol_level2.netset",
    "blocklist_de": "https://lists.blocklist.de/lists/all.txt",
    "feodotracker": "https://feodotracker.abuse.ch/downloads/ipblocklist.txt",
    "cins_army": "http://cinsscore.com/list/ci-badguys.txt",
}

VT_SLEEP_SECONDS = 16          # keeps you under VirusTotal's 4 req/min
ABUSEIPDB_SLEEP_SECONDS = 0.3  # well under AbuseIPDB's 5 req/sec
GREYNOISE_SLEEP_SECONDS = 0.5
INTERNETDB_CONCURRENCY = 20    # polite concurrency, InternetDB can handle far more

CHECKPOINT_SAVE_EVERY = 50


# --------------------------------------------------------------------------
# Quota tracking (resets automatically at midnight local time)
# --------------------------------------------------------------------------

def load_quota():
    today = str(date.today())
    if QUOTA_FILE.exists():
        data = json.loads(QUOTA_FILE.read_text())
        if data.get("date") == today:
            return data
    return {"date": today, "used": {k: 0 for k in DAILY_QUOTAS}}


def save_quota(quota):
    QUOTA_FILE.write_text(json.dumps(quota, indent=2))


def quota_remaining(quota, key):
    return DAILY_QUOTAS[key] - quota["used"].get(key, 0)


def quota_consume(quota, key, n=1):
    quota["used"][key] = quota["used"].get(key, 0) + n


# --------------------------------------------------------------------------
# Checkpoint (per-IP results, so re-runs don't repeat work)
# --------------------------------------------------------------------------

def load_checkpoint():
    if CHECKPOINT_FILE.exists():
        return json.loads(CHECKPOINT_FILE.read_text())
    return {}


def save_checkpoint(checkpoint):
    CHECKPOINT_FILE.write_text(json.dumps(checkpoint, indent=2))


# --------------------------------------------------------------------------
# Tier 0: local static blocklists
# --------------------------------------------------------------------------

def download_blocklists():
    BLOCKLIST_CACHE_DIR.mkdir(exist_ok=True)
    for name, url in BLOCKLIST_SOURCES.items():
        cache_file = BLOCKLIST_CACHE_DIR / f"{name}.txt"
        if cache_file.exists():
            age_hours = (time.time() - cache_file.stat().st_mtime) / 3600
            if age_hours < BLOCKLIST_MAX_AGE_HOURS:
                continue
        try:
            resp = HTTP.get(url, timeout=30, headers={"User-Agent": "ip-recon-script/1.0"})
            resp.raise_for_status()
            cache_file.write_text(resp.text)
            print(f"  [blocklist] refreshed {name} ({len(resp.text.splitlines())} lines)")
        except Exception as exc:
            if cache_file.exists():
                print(f"  [blocklist] {name} refresh failed ({exc}); using cached copy")
            else:
                print(f"  [blocklist] {name} failed and no cache available ({exc}); skipping")


def build_blocklist_index():
    """Returns a sorted list of (start_int, end_int, source) for bisect lookups."""
    ranges = []
    for name in BLOCKLIST_SOURCES:
        cache_file = BLOCKLIST_CACHE_DIR / f"{name}.txt"
        if not cache_file.exists():
            continue
        for line in cache_file.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith(("#", ";")):
                continue
            token = line.split()[0]  # some lists have trailing comments/columns
            try:
                net = ipaddress.ip_network(token, strict=False)
            except ValueError:
                continue
            if net.version != 4:
                continue  # keep it simple; extend to v6 if you need it
            ranges.append((int(net.network_address), int(net.broadcast_address), name))
    ranges.sort(key=lambda r: r[0])
    starts = [r[0] for r in ranges]
    return ranges, starts


def check_local_blocklists(ip, ranges, starts):
    try:
        addr = ipaddress.ip_address(ip)
    except ValueError:
        return []
    if addr.version != 4:
        # The static lists wired up in BLOCKLIST_SOURCES are IPv4-only feeds.
        # Returning immediately avoids walking the whole IPv4 range table
        # for every IPv6 address (which would never match anyway).
        return []
    ip_int = int(addr)
    matches = []
    idx = bisect.bisect_right(starts, ip_int) - 1
    # Walk backward since overlapping ranges from different lists are possible
    i = idx
    while i >= 0 and ranges[i][0] <= ip_int:
        start, end, source = ranges[i]
        if start <= ip_int <= end:
            matches.append(source)
        i -= 1
        if idx - i > 5000:  # safety valve against pathological overlap
            break
    return matches


# --------------------------------------------------------------------------
# Tier 0.5: AbuseIPDB blacklist pull (cheap, covers up to 10k IPs per call)
# --------------------------------------------------------------------------

def pull_abuseipdb_blacklist(quota):
    if not ABUSEIPDB_API_KEY:
        return {}
    if quota_remaining(quota, "abuseipdb_blacklist") <= 0:
        print("  [abuseipdb-blacklist] daily quota used up, skipping")
        return {}
    try:
        resp = HTTP.get(
            "https://api.abuseipdb.com/api/v2/blacklist",
            headers={"Key": ABUSEIPDB_API_KEY, "Accept": "application/json"},
            params={"confidenceMinimum": 75, "limit": 10000},
            timeout=30,
        )
        quota_consume(quota, "abuseipdb_blacklist")
        resp.raise_for_status()
        data = resp.json().get("data", [])
        result = {row["ipAddress"]: row["abuseConfidenceScore"] for row in data}
        print(f"  [abuseipdb-blacklist] pulled {len(result)} known-bad IPs")
        return result
    except Exception as exc:
        print(f"  [abuseipdb-blacklist] failed: {exc}")
        return {}


# --------------------------------------------------------------------------
# Tier 1: Shodan InternetDB (no key required)
# --------------------------------------------------------------------------

def check_internetdb(ip):
    try:
        resp = HTTP.get(f"https://internetdb.shodan.io/{ip}", timeout=10)
        if resp.status_code == 404:
            return {"found": False}
        resp.raise_for_status()
        data = resp.json()
        return {
            "found": True,
            "ports": data.get("ports", []),
            "tags": data.get("tags", []),
            "vulns": data.get("vulns", []),
            "hostnames": data.get("hostnames", []),
        }
    except Exception:
        return {"error": True}


def run_internetdb_tier(ips, checkpoint):
    todo = [ip for ip in ips if "internetdb" not in checkpoint.get(ip, {})]
    if not todo:
        return
    print(f"[Tier 1] Shodan InternetDB: {len(todo)} IPs to check")
    with ThreadPoolExecutor(max_workers=INTERNETDB_CONCURRENCY) as pool:
        futures = {pool.submit(check_internetdb, ip): ip for ip in todo}
        count = 0
        for fut in tqdm(as_completed(futures), total=len(futures), desc="InternetDB"):
            ip = futures[fut]
            checkpoint.setdefault(ip, {})["internetdb"] = fut.result()
            count += 1
            if count % CHECKPOINT_SAVE_EVERY == 0:
                save_checkpoint(checkpoint)
    save_checkpoint(checkpoint)


# --------------------------------------------------------------------------
# Tier 2: individual API lookups (rate-limited, quota-tracked, resumable)
# --------------------------------------------------------------------------

def check_abuseipdb(ip):
    resp = HTTP.get(
        "https://api.abuseipdb.com/api/v2/check",
        headers={"Key": ABUSEIPDB_API_KEY, "Accept": "application/json"},
        params={"ipAddress": ip, "maxAgeInDays": 90},
        timeout=15,
    )
    resp.raise_for_status()
    d = resp.json()["data"]
    return {
        "score": d.get("abuseConfidenceScore"),
        "reports": d.get("totalReports"),
        "country": d.get("countryCode"),
        "isp": d.get("isp"),
        "usage_type": d.get("usageType"),
        "is_tor": d.get("isTor"),
    }


def check_virustotal(ip):
    resp = HTTP.get(
        f"https://www.virustotal.com/api/v3/ip_addresses/{ip}",
        headers={"x-apikey": VT_API_KEY},
        timeout=15,
    )
    resp.raise_for_status()
    stats = resp.json()["data"]["attributes"]["last_analysis_stats"]
    return {
        "malicious": stats.get("malicious", 0),
        "suspicious": stats.get("suspicious", 0),
        "harmless": stats.get("harmless", 0),
        "undetected": stats.get("undetected", 0),
    }


def check_greynoise(ip):
    resp = HTTP.get(
        f"https://api.greynoise.io/v3/community/{ip}",
        headers={"key": GREYNOISE_API_KEY, "Accept": "application/json"},
        timeout=15,
    )
    if resp.status_code == 404:
        return {"noise": False, "classification": "unknown"}
    resp.raise_for_status()
    d = resp.json()
    return {
        "noise": d.get("noise"),
        "riot": d.get("riot"),
        "classification": d.get("classification"),
        "name": d.get("name"),
    }


def run_individual_api_tier(ips, checkpoint, quota, known_bad_from_blacklist):
    """
    Only spends per-IP API budget on IPs that Tier 0 / Tier 0.5 didn't
    already resolve, so the limited daily quotas go furthest.
    """
    needs_lookup = [
        ip for ip in ips
        if not checkpoint.get(ip, {}).get("blocklist_hits")
        and ip not in known_bad_from_blacklist
    ]

    # --- AbuseIPDB individual checks ---
    if ABUSEIPDB_API_KEY:
        todo = [ip for ip in needs_lookup if "abuseipdb" not in checkpoint.get(ip, {})]
        remaining = quota_remaining(quota, "abuseipdb_check")
        batch = todo[:remaining]
        if batch:
            print(f"[Tier 2] AbuseIPDB individual check: {len(batch)} IPs "
                  f"({len(todo) - len(batch)} left for a future run)")
        for i, ip in enumerate(tqdm(batch, desc="AbuseIPDB")):
            try:
                checkpoint.setdefault(ip, {})["abuseipdb"] = check_abuseipdb(ip)
            except Exception as exc:
                checkpoint.setdefault(ip, {})["abuseipdb"] = {"error": str(exc)}
            quota_consume(quota, "abuseipdb_check")
            if i % CHECKPOINT_SAVE_EVERY == 0:
                save_checkpoint(checkpoint)
                save_quota(quota)
            time.sleep(ABUSEIPDB_SLEEP_SECONDS)
        save_checkpoint(checkpoint)
        save_quota(quota)
    else:
        print("[Tier 2] Skipping AbuseIPDB (no ABUSEIPDB_API_KEY set)")

    # --- VirusTotal individual checks ---
    if VT_API_KEY:
        todo = [ip for ip in needs_lookup if "virustotal" not in checkpoint.get(ip, {})]
        remaining = quota_remaining(quota, "virustotal")
        batch = todo[:remaining]
        if batch:
            print(f"[Tier 2] VirusTotal individual check: {len(batch)} IPs "
                  f"({len(todo) - len(batch)} left for a future run, "
                  f"~{len(batch) * VT_SLEEP_SECONDS / 60:.0f} min)")
        for i, ip in enumerate(tqdm(batch, desc="VirusTotal")):
            try:
                checkpoint.setdefault(ip, {})["virustotal"] = check_virustotal(ip)
            except Exception as exc:
                checkpoint.setdefault(ip, {})["virustotal"] = {"error": str(exc)}
            quota_consume(quota, "virustotal")
            if i % CHECKPOINT_SAVE_EVERY == 0:
                save_checkpoint(checkpoint)
                save_quota(quota)
            time.sleep(VT_SLEEP_SECONDS)
        save_checkpoint(checkpoint)
        save_quota(quota)
    else:
        print("[Tier 2] Skipping VirusTotal (no VT_API_KEY set)")

    # --- GreyNoise individual checks (optional) ---
    if GREYNOISE_API_KEY:
        todo = [ip for ip in needs_lookup if "greynoise" not in checkpoint.get(ip, {})]
        remaining = quota_remaining(quota, "greynoise")
        batch = todo[:remaining]
        if batch:
            print(f"[Tier 2] GreyNoise individual check: {len(batch)} IPs "
                  f"({len(todo) - len(batch)} left for a future run)")
        for i, ip in enumerate(tqdm(batch, desc="GreyNoise")):
            try:
                checkpoint.setdefault(ip, {})["greynoise"] = check_greynoise(ip)
            except Exception as exc:
                checkpoint.setdefault(ip, {})["greynoise"] = {"error": str(exc)}
            quota_consume(quota, "greynoise")
            if i % CHECKPOINT_SAVE_EVERY == 0:
                save_checkpoint(checkpoint)
                save_quota(quota)
            time.sleep(GREYNOISE_SLEEP_SECONDS)
        save_checkpoint(checkpoint)
        save_quota(quota)
    else:
        print("[Tier 2] Skipping GreyNoise (no GREYNOISE_API_KEY set)")


# --------------------------------------------------------------------------
# Verdict logic
# --------------------------------------------------------------------------

def compute_verdict(row):
    blocklist_hits = row.get("blocklist_hits") or []
    blacklist_score = row.get("abuseipdb_blacklist_score")
    abuseipdb = row.get("abuseipdb") or {}
    vt = row.get("virustotal") or {}
    greynoise = row.get("greynoise") or {}

    if blocklist_hits or (blacklist_score is not None and blacklist_score >= 90):
        return "Malicious"
    if abuseipdb.get("score") is not None and abuseipdb["score"] >= 75:
        return "Malicious"
    if vt.get("malicious", 0) >= 3:
        return "Malicious"
    if greynoise.get("classification") == "malicious":
        return "Malicious"

    if (blacklist_score is not None and blacklist_score >= 25) \
            or (abuseipdb.get("score") is not None and 25 <= abuseipdb["score"] < 75) \
            or (0 < vt.get("malicious", 0) < 3) \
            or vt.get("suspicious", 0) > 0 \
            or greynoise.get("classification") == "suspicious":
        return "Suspicious"

    have_any_data = any([
        blacklist_score is not None,
        abuseipdb.get("score") is not None,
        "malicious" in vt,
        greynoise.get("classification") not in (None, "unknown"),
    ])
    if have_any_data:
        return "Not Malicious"
    return "Unknown / Not Checked Yet"


# --------------------------------------------------------------------------
# Excel export
# --------------------------------------------------------------------------

VERDICT_COLORS = {
    "Malicious": "FFC7CE",
    "Suspicious": "FFEB9C",
    "Not Malicious": "C6EFCE",
    "Unknown / Not Checked Yet": "D9D9D9",
}


def export_excel(ips, checkpoint, blacklist_map, output_path):
    rows = []
    for ip in ips:
        rec = dict(checkpoint.get(ip, {}))
        rec["abuseipdb_blacklist_score"] = blacklist_map.get(ip)
        verdict = compute_verdict(rec)

        internetdb = rec.get("internetdb") or {}
        abuseipdb = rec.get("abuseipdb") or {}
        vt = rec.get("virustotal") or {}
        greynoise = rec.get("greynoise") or {}

        rows.append({
            "IP": ip,
            "IP Version": "IPv6" if ":" in ip else "IPv4",
            "Verdict": verdict,
            "Blocklist Sources": ", ".join(rec.get("blocklist_hits") or []),
            "AbuseIPDB Blacklist Score": rec.get("abuseipdb_blacklist_score"),
            "AbuseIPDB Score": abuseipdb.get("score"),
            "AbuseIPDB Reports": abuseipdb.get("reports"),
            "AbuseIPDB Country": abuseipdb.get("country"),
            "AbuseIPDB ISP": abuseipdb.get("isp"),
            "AbuseIPDB Usage Type": abuseipdb.get("usage_type"),
            "AbuseIPDB Is Tor": abuseipdb.get("is_tor"),
            "VT Malicious": vt.get("malicious"),
            "VT Suspicious": vt.get("suspicious"),
            "VT Harmless": vt.get("harmless"),
            "GreyNoise Classification": greynoise.get("classification"),
            "GreyNoise Name": greynoise.get("name"),
            "Shodan Open Ports": ", ".join(str(p) for p in internetdb.get("ports", []) or []),
            "Shodan Known CVEs": ", ".join(internetdb.get("vulns", []) or []),
            "Shodan Tags": ", ".join(internetdb.get("tags", []) or []),
        })

    df = pd.DataFrame(rows)
    df.to_excel(output_path, index=False, sheet_name="IP Reputation")

    # Styling pass
    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb["IP Reputation"]
    ws.freeze_panes = "A2"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    verdict_col_idx = df.columns.get_loc("Verdict") + 1
    for row_idx in range(2, ws.max_row + 1):
        verdict = ws.cell(row=row_idx, column=verdict_col_idx).value
        color = VERDICT_COLORS.get(verdict)
        if color:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    wb.save(output_path)


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def load_ip_list(path):
    ips = []
    seen = set()
    if not Path(path).exists():
        print(f"\nERROR: input file not found: {path}")
        print("Check the INPUT_FILE_PATH in the CONFIG box at the top of this "
              "script (or the --input argument if you passed one).")
        sys.exit(1)
    with open(path) as f:
        for line in f:
            token = line.strip()
            if not token:
                continue
            try:
                ipaddress.ip_address(token)
            except ValueError:
                print(f"  [skip] not a valid IP: {token}")
                continue
            if token not in seen:
                seen.add(token)
                ips.append(token)
    return ips


def apply_blocklist_tier(ips, checkpoint):
    print("[Tier 0] Refreshing local blocklists...")
    download_blocklists()
    ranges, starts = build_blocklist_index()
    print(f"  loaded {len(ranges)} blocklist CIDR ranges")
    todo = [ip for ip in ips if "blocklist_hits" not in checkpoint.get(ip, {})]
    for ip in tqdm(todo, desc="Local blocklists"):
        checkpoint.setdefault(ip, {})["blocklist_hits"] = check_local_blocklists(ip, ranges, starts)
    save_checkpoint(checkpoint)


def main():
    parser = argparse.ArgumentParser(description="Bulk IP reputation checker (free-tier friendly)")
    parser.add_argument("--input", default=INPUT_FILE_PATH,
                         help=f"Path to text file, one IP per line (default from CONFIG box: {INPUT_FILE_PATH})")
    parser.add_argument("--output", default=OUTPUT_FILE_PATH,
                         help=f"Output Excel path (default from CONFIG box: {OUTPUT_FILE_PATH})")
    parser.add_argument("--skip-internetdb", action="store_true")
    parser.add_argument("--skip-individual-apis", action="store_true",
                         help="Only run Tier 0/0.5/1 (free, unlimited, fast pass)")
    args = parser.parse_args()

    print("Configured vendors:")
    print(f"  AbuseIPDB : {'ENABLED' if ABUSEIPDB_API_KEY else 'skipped (no key in CONFIG box or env)'}")
    print(f"  VirusTotal: {'ENABLED' if VT_API_KEY else 'skipped (no key in CONFIG box or env)'}")
    print(f"  GreyNoise : {'ENABLED' if GREYNOISE_API_KEY else 'skipped (no key in CONFIG box or env)'}")
    print()

    print(f"Loading IPs from {args.input} ...")
    ips = load_ip_list(args.input)
    print(f"  {len(ips)} unique valid IPs loaded")

    checkpoint = load_checkpoint()
    quota = load_quota()

    # Tier 0: local blocklists (free, instant, all IPs)
    apply_blocklist_tier(ips, checkpoint)

    # Tier 0.5: AbuseIPDB blacklist pull (cheap, covers thousands at once)
    blacklist_map = pull_abuseipdb_blacklist(quota)
    save_quota(quota)

    # Tier 1: Shodan InternetDB (free, no key, all IPs)
    if not args.skip_internetdb:
        run_internetdb_tier(ips, checkpoint)

    # Tier 2: individual rate-limited API lookups on whatever is left unresolved
    if not args.skip_individual_apis:
        run_individual_api_tier(ips, checkpoint, quota, blacklist_map)

    # Export
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    print(f"Writing report to {output_path} ...")
    export_excel(ips, checkpoint, blacklist_map, str(output_path))
    print("Done.")

    # Summary
    verdicts = {}
    for ip in ips:
        rec = dict(checkpoint.get(ip, {}))
        rec["abuseipdb_blacklist_score"] = blacklist_map.get(ip)
        v = compute_verdict(rec)
        verdicts[v] = verdicts.get(v, 0) + 1
    print("\nSummary:")
    for v, c in sorted(verdicts.items(), key=lambda x: -x[1]):
        print(f"  {v}: {c}")

    unknown = verdicts.get("Unknown / Not Checked Yet", 0)
    if unknown:
        print(f"\n{unknown} IPs still unresolved (daily API quota reached or no keys set).")
        print("Re-run this exact command tomorrow (or later today for VirusTotal, "
              "which resets on a rolling basis) to continue -- progress is saved "
              "in checkpoint.json / quota_tracker.json.")


if __name__ == "__main__":
    main()
